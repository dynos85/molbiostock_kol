import pandas as pd
import streamlit as st
from io import BytesIO
import datetime

def export_data(db, data_type):
    """Export data to Excel with formatting"""

    if data_type == "stock":
        df = db.get_current_stock()
        filename = "current_stock"
    elif data_type == "transactions":
        df = db.get_all_transactions()
        filename = "transactions"
    else:
        st.error("Invalid export type")
        return None

    # Create Excel writer object
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=data_type.title())

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[data_type.title()]

        # Apply formatting to headers
        from openpyxl.styles import Font, PatternFill
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4361EE', end_color='4361EE', fill_type='solid')

        # Auto-adjust columns
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Generate filename with timestamp
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    final_filename = f"{filename}_{timestamp}.xlsx"

    return output.getvalue(), final_filename

def get_csv_download_link(df, filename):
    """Generate CSV download link"""
    csv = df.to_csv(index=False)
    b64 = base64.b64encode(csv.encode()).decode()
    href = f'<a href="data:file/csv;base64,{b64}" download="{filename}">Download CSV</a>'
    return href

import base64